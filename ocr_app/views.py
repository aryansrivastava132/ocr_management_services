"""
ocr_app/views.py
"""
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.shortcuts import render, get_object_or_404
from django.utils.decorators import method_decorator

from .models import ScanRecord, DateChart
from .ocr_utils import extract_text, parse_fields, validate


# ── Main page ──────────────────────────────────────────────────────────────
def index(request):
    return render(request, 'ocr_app/index.html')


# ── Scan endpoint ──────────────────────────────────────────────────────────
@csrf_exempt
@require_POST
def scan(request):
    if 'photo' not in request.FILES:
        return JsonResponse({'error': 'No photo uploaded'}, status=400)

    photo = request.FILES['photo']
    allowed = {'image/jpeg', 'image/png', 'image/webp', 'image/bmp'}
    if photo.content_type not in allowed:
        return JsonResponse({'error': 'Unsupported file type'}, status=400)

    # Save record first (to get file path)
    record = ScanRecord(image=photo)
    record.save()

    try:
        raw_text = extract_text(record.image.path)
        fields   = parse_fields(raw_text)
        result   = validate(fields)

        record.mfd      = fields['mfd']
        record.exp      = fields['exp']
        record.batch    = fields['batch']
        record.raw_text = raw_text
        record.result   = result
        record.save()

    except Exception as e:
        record.result   = 'FAIL'
        record.raw_text = f'OCR Error: {str(e)}'
        record.save()
        return JsonResponse({
            'scan_id'  : record.pk,
            'result'   : 'FAIL',
            'fields'   : {'mfd': '', 'exp': '', 'batch': ''},
            'raw_text' : record.raw_text,
            'error'    : str(e),
        })

    return JsonResponse({
        'scan_id'    : record.pk,
        'result'     : result,
        'fields'     : fields,
        'raw_text'   : raw_text[:500],
        'image_url'  : request.build_absolute_uri(record.image.url),
        'scanned_at' : record.scanned_at.strftime('%d %b %Y, %I:%M %p'),
    })


# ── History API ────────────────────────────────────────────────────────────
@require_GET
def history_api(request):
    scans = ScanRecord.objects.all()[:50]
    data  = [{
        'id'        : s.pk,
        'mfd'       : s.mfd,
        'exp'       : s.exp,
        'batch'     : s.batch,
        'result'    : s.result,
        'scanned_at': s.scanned_at.strftime('%d %b %Y, %I:%M %p'),
        'image_url' : request.build_absolute_uri(s.image.url) if s.image else '',
    } for s in scans]
    return JsonResponse({'scans': data})


# ── Date chart CRUD ────────────────────────────────────────────────────────
@csrf_exempt
def date_chart_api(request):
    if request.method == 'GET':
        codes = list(DateChart.objects.values('id', 'code', 'valid_date', 'product_type'))
        return JsonResponse({'codes': codes})

    if request.method == 'POST':
        body = json.loads(request.body)
        obj, created = DateChart.objects.update_or_create(
            code=body['code'].upper().strip(),
            defaults={
                'valid_date'  : body.get('valid_date', ''),
                'product_type': body.get('product_type', ''),
            }
        )
        return JsonResponse({'id': obj.pk, 'created': created})

    if request.method == 'DELETE':
        body = json.loads(request.body)
        DateChart.objects.filter(pk=body['id']).delete()
        return JsonResponse({'deleted': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ── Excel report download ──────────────────────────────────────────────────
@require_GET
def download_report(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scan Report"

    headers    = ['ID', 'MFD', 'EXP', 'Batch No', 'Result', 'Scanned At']
    hdr_fill   = PatternFill("solid", fgColor="1a3c5e")
    pass_fill  = PatternFill("solid", fgColor="22c55e")
    fail_fill  = PatternFill("solid", fgColor="ef4444")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(color="FFFFFF", bold=True, size=11)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for row_i, s in enumerate(ScanRecord.objects.all(), 2):
        ws.cell(row=row_i, column=1, value=s.pk)
        ws.cell(row=row_i, column=2, value=s.mfd)
        ws.cell(row=row_i, column=3, value=s.exp)
        ws.cell(row=row_i, column=4, value=s.batch)
        rc = ws.cell(row=row_i, column=5, value=s.result)
        rc.fill = pass_fill if s.result == 'PASS' else fail_fill
        rc.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row_i, column=6, value=s.scanned_at.strftime('%d %b %Y %H:%M'))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"scan_report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
